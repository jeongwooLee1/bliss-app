import sys, re, os
sys.stdout.reconfigure(encoding='utf-8')
import oracledb
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
from collections import defaultdict

oracledb.init_oracle_client(lib_dir=r"C:\Users\TP005\instantclient_21_15")
conn = oracledb.connect(user="housewaxing", password="oracle", dsn="googlea.withbiz.co.kr:5063/ORA11GHW")
cur = conn.cursor()

BRANCH_MAP = {9:"강남점", 10:"왕십리점", 11:"홍대점", 20:"잠실점(구)", 21:"해운대점",
              30:"용산점(구)", 40:"잠실점", 50:"마곡점", 60:"용산점", 70:"위례점", 71:"천호점"}

today = datetime.now()
hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
hdr_font = Font(bold=True, color="FFFFFF", size=11)

def write_headers(ws, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')

# ─── 1. 패키지(다회권) ───
print("패키지 데이터 추출 중...")
cur.execute("""
    SELECT o.MEMNO, s.FIRSTNAME, s.OPKGC, o.ORDERDATE, o.BRANCHNO
    FROM ORDERS o
    JOIN ORDERDETAIL d ON d.ORDERNO = o.NO
    JOIN SERVICE s ON d.SRVNO = s.NO
    WHERE s.SRVDIV = 'PACK'
    ORDER BY o.MEMNO, o.ORDERDATE
""")
rows = cur.fetchall()

def normalize_pkg(name):
    n = name.replace("패키지권","").replace("패키지","").strip()
    n = re.sub(r'\d+$', '', n).strip()
    return n

pkg_data = defaultdict(lambda: {"bought":0, "used":0, "last_buy":None, "last_use":None, "last_activity":None, "branch":None})
for memno, svcname, opkgc, orderdate, branchno in rows:
    key = (memno, normalize_pkg(svcname))
    if opkgc and opkgc > 0:
        pkg_data[key]["bought"] += opkgc
        if not pkg_data[key]["last_buy"] or orderdate > pkg_data[key]["last_buy"]:
            pkg_data[key]["last_buy"] = orderdate
            pkg_data[key]["branch"] = branchno
    else:
        pkg_data[key]["used"] += 1
        if not pkg_data[key]["last_use"] or orderdate > pkg_data[key]["last_use"]:
            pkg_data[key]["last_use"] = orderdate
    if not pkg_data[key]["last_activity"] or orderdate > pkg_data[key]["last_activity"]:
        pkg_data[key]["last_activity"] = orderdate
        if not pkg_data[key]["branch"]:
            pkg_data[key]["branch"] = branchno

pkg_remaining = []
for (memno, pkg_type), d in pkg_data.items():
    remain = d["bought"] - d["used"]
    if remain > 0:
        pkg_remaining.append((memno, pkg_type, d["bought"], d["used"], remain,
                              d["last_buy"], d["last_use"], d["last_activity"], d["branch"]))

# 고객 정보 조회
all_memnos = list(set(r[0] for r in pkg_remaining))
mem_map = {}
for i in range(0, len(all_memnos), 500):
    batch = all_memnos[i:i+500]
    cur.execute(f"SELECT NO, NAME, TEL, SEXDIV FROM MEMBER WHERE NO IN ({','.join(str(m) for m in batch)})")
    for r in cur.fetchall():
        mem_map[r[0]] = {"name": r[1], "tel": r[2], "sex": r[3]}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "패키지(다회권)"

headers = ["고객번호","이름","전화번호","성별","패키지종류","구매횟수","사용횟수","잔여횟수",
           "마지막구매일","마지막사용일","마지막활동일","만료일","만료여부","지점"]
write_headers(ws, headers)

pkg_remaining.sort(key=lambda x: (x[8] or 0, x[1], -(x[4])))
for i, (memno, pkg_type, bought, used, remain, last_buy, last_use, last_act, branch) in enumerate(pkg_remaining, 2):
    m = mem_map.get(memno, {})
    # 만료기준: 사용이력 있으면 마지막사용일, 없으면 마지막구매일 기준 1년
    base_date = last_use if last_use else last_buy
    expire_date = (base_date + timedelta(days=365)) if base_date else None
    is_expired = expire_date < today if expire_date else True

    ws.cell(row=i, column=1, value=memno)
    ws.cell(row=i, column=2, value=m.get("name",""))
    ws.cell(row=i, column=3, value=m.get("tel",""))
    ws.cell(row=i, column=4, value=m.get("sex",""))
    ws.cell(row=i, column=5, value=pkg_type)
    ws.cell(row=i, column=6, value=bought)
    ws.cell(row=i, column=7, value=used)
    ws.cell(row=i, column=8, value=remain)
    ws.cell(row=i, column=9, value=last_buy.strftime("%Y-%m-%d") if last_buy else "")
    ws.cell(row=i, column=10, value=last_use.strftime("%Y-%m-%d") if last_use else "")
    ws.cell(row=i, column=11, value=last_act.strftime("%Y-%m-%d") if last_act else "")
    ws.cell(row=i, column=12, value=expire_date.strftime("%Y-%m-%d") if expire_date else "")
    ws.cell(row=i, column=13, value="만료" if is_expired else "유효")
    ws.cell(row=i, column=14, value=BRANCH_MAP.get(branch, str(branch) if branch else ""))

    if is_expired:
        for col in range(1, 15):
            ws.cell(row=i, column=col).font = Font(color="999999")

for col in range(1, 15):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14

# ─── 2. 선불금(다담권) ───
print("선불금 데이터 추출 중...")
ws2 = wb.create_sheet("선불금(다담권)")

cur.execute("""
    SELECT o.MEMNO, m.NAME, m.TEL, m.SEXDIV, o.ORDERDATE, o.NOTE, o.BRANCHNO
    FROM ORDERS o
    JOIN MEMBER m ON o.MEMNO = m.NO
    WHERE o.NOTE LIKE '%잔여%' OR o.NOTE LIKE '%다담%' OR o.NOTE LIKE '%선불%'
    ORDER BY o.MEMNO, o.ORDERDATE DESC
""")
prepaid_rows = cur.fetchall()

seen = set()
prepaid_results = []
for memno, name, tel, sex, dt, note, branch in prepaid_rows:
    if memno in seen:
        continue
    seen.add(memno)
    matches = re.findall(r'잔여[금액:\s]*([0-9,.]+)', note or "")
    if matches:
        amt_str = matches[-1].replace(",", "").replace(".", "").strip()
        try:
            amt = int(amt_str)
            if amt > 0:
                prepaid_results.append((memno, name, tel, sex, dt, amt, branch, (note or "")[:100]))
        except:
            pass

headers2 = ["고객번호","이름","전화번호","성별","마지막사용일","잔여금액","만료일","만료여부","지점","메모"]
write_headers(ws2, headers2)

prepaid_results.sort(key=lambda x: -(x[5]))
for i, (memno, name, tel, sex, dt, amt, branch, note) in enumerate(prepaid_results, 2):
    expire_date = (dt + timedelta(days=365)) if dt else None
    is_expired = expire_date < today if expire_date else True

    ws2.cell(row=i, column=1, value=memno)
    ws2.cell(row=i, column=2, value=name)
    ws2.cell(row=i, column=3, value=tel)
    ws2.cell(row=i, column=4, value=sex or "")
    ws2.cell(row=i, column=5, value=dt.strftime("%Y-%m-%d") if dt else "")
    ws2.cell(row=i, column=6, value=amt)
    ws2.cell(row=i, column=7, value=expire_date.strftime("%Y-%m-%d") if expire_date else "")
    ws2.cell(row=i, column=8, value="만료" if is_expired else "유효")
    ws2.cell(row=i, column=9, value=BRANCH_MAP.get(branch, str(branch) if branch else ""))
    ws2.cell(row=i, column=10, value=note)

    if is_expired:
        for col in range(1, 11):
            ws2.cell(row=i, column=col).font = Font(color="999999")

for col in range(1, 11):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
ws2.column_dimensions['J'].width = 60

# ─── 3. 포인트(적립금) ───
print("포인트 데이터 추출 중...")
ws3 = wb.create_sheet("포인트(적립금)")

cur.execute("""
    SELECT p.MEMNO, m.NAME, m.TEL, m.SEXDIV,
           SUM(p.POINT) as balance,
           MAX(p.MODDATE) as last_date,
           COUNT(*) as txn_cnt,
           p.BRANCHDIV
    FROM POINT p
    JOIN MEMBER m ON p.MEMNO = m.NO
    GROUP BY p.MEMNO, m.NAME, m.TEL, m.SEXDIV, p.BRANCHDIV
    HAVING SUM(p.POINT) > 0
    ORDER BY SUM(p.POINT) DESC
""")
point_rows = cur.fetchall()

headers3 = ["고객번호","이름","전화번호","성별","잔액","마지막거래일","만료일","만료여부","거래건수","지점"]
write_headers(ws3, headers3)

for i, (memno, name, tel, sex, balance, last_dt, txn_cnt, branchdiv) in enumerate(point_rows, 2):
    expire_date = (last_dt + timedelta(days=365)) if last_dt else None
    is_expired = expire_date < today if expire_date else True

    ws3.cell(row=i, column=1, value=memno)
    ws3.cell(row=i, column=2, value=name)
    ws3.cell(row=i, column=3, value=tel)
    ws3.cell(row=i, column=4, value=sex or "")
    ws3.cell(row=i, column=5, value=balance)
    ws3.cell(row=i, column=6, value=last_dt.strftime("%Y-%m-%d") if last_dt else "")
    ws3.cell(row=i, column=7, value=expire_date.strftime("%Y-%m-%d") if expire_date else "")
    ws3.cell(row=i, column=8, value="만료" if is_expired else "유효")
    ws3.cell(row=i, column=9, value=txn_cnt)
    ws3.cell(row=i, column=10, value=branchdiv or "")

    if is_expired:
        for col in range(1, 11):
            ws3.cell(row=i, column=col).font = Font(color="999999")

for col in range(1, 11):
    ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14

# 저장
outpath = r"C:\Users\TP005\Desktop\오라클_패키지_선불금_포인트_v2.xlsx"
wb.save(outpath)
print(f"\n저장 완료: {outpath}")

# 통계
valid_pkg = sum(1 for r in pkg_remaining if r[7] and (r[7] + timedelta(days=365)) >= today)
valid_prepaid = sum(1 for r in prepaid_results if r[4] and (r[4] + timedelta(days=365)) >= today)
valid_point = sum(1 for r in point_rows if r[5] and (r[5] + timedelta(days=365)) >= today)

print(f"\n=== 통계 ===")
print(f"패키지: 전체 {len(pkg_remaining)}건, 유효 {valid_pkg}건, 만료 {len(pkg_remaining)-valid_pkg}건")
print(f"선불금: 전체 {len(prepaid_results)}명, 유효 {valid_prepaid}명, 만료 {len(prepaid_results)-valid_prepaid}명")
print(f"포인트: 전체 {len(point_rows)}명, 유효 {valid_point}명, 만료 {len(point_rows)-valid_point}명")

cur.close()
conn.close()
