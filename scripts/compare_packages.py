import sys, re, json
sys.stdout.reconfigure(encoding='utf-8')
import oracledb
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
from collections import defaultdict

oracledb.init_oracle_client(lib_dir=r"C:\Users\TP005\instantclient_21_15")
conn = oracledb.connect(user="housewaxing", password="oracle", dsn="googlea.withbiz.co.kr:5063/ORA11GHW")
cur = conn.cursor()

import requests
SB_URL = "https://dpftlrsuqxqqeouwbfjd.supabase.co"
SB_KEY = "sb_publishable_3H-KTP0MoV_KuY74ocbefw_3Ze5xBJj"
SB_HEADERS = {"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}"}

BRANCH_MAP = {9:"강남점", 10:"왕십리점", 11:"홍대점", 20:"잠실점(구)", 21:"해운대점",
              30:"용산점(구)", 40:"잠실점", 50:"마곡점", 60:"용산점", 70:"위례점", 71:"천호점"}
today = datetime.now()

# ═══════════════════════════════════════════
# 1. 오라클 ORDERDETAIL 기반 패키지
# ═══════════════════════════════════════════
print("1. 오라클 패키지 추출 중...")
cur.execute("""
    SELECT o.MEMNO, s.FIRSTNAME, s.OPKGC, o.ORDERDATE, o.BRANCHNO
    FROM ORDERS o
    JOIN ORDERDETAIL d ON d.ORDERNO = o.NO
    JOIN SERVICE s ON d.SRVNO = s.NO
    WHERE s.SRVDIV = 'PACK'
    ORDER BY o.MEMNO, o.ORDERDATE
""")
rows = cur.fetchall()

def normalize_pkg(name):
    n = name.replace("패키지권","").replace("패키지","").strip()
    n = re.sub(r'\d+$', '', n).strip()
    return n

pkg_events = defaultdict(list)
for memno, svcname, opkgc, orderdate, branchno in rows:
    key = (memno, normalize_pkg(svcname))
    if opkgc and opkgc > 0:
        pkg_events[key].append((orderdate, "buy", opkgc, branchno))
    else:
        pkg_events[key].append((orderdate, "use", 1, branchno))

oracle_pkgs = {}  # key=(memno, pkg_type) → {remain, last_buy, last_use, branch}
for (memno, pkg_type), evts in pkg_events.items():
    evts.sort(key=lambda x: x[0])
    total_bought = sum(e[2] for e in evts if e[1] == "buy")
    total_used = sum(e[2] for e in evts if e[1] == "use")
    remain = total_bought - total_used
    if remain <= 0:
        continue
    buy_events = [e for e in evts if e[1] == "buy"]
    use_events = [e for e in evts if e[1] == "use"]
    last_buy = buy_events[-1][0] if buy_events else None
    last_use = use_events[-1][0] if use_events else None
    branch = buy_events[-1][3] if buy_events else None
    # 2년 초과 건 제외
    base_date = last_use if last_use else last_buy
    if base_date and (today - base_date).days > 730:
        continue
    oracle_pkgs[(memno, pkg_type)] = {
        "remain": remain, "bought": total_bought, "used": total_used,
        "last_buy": last_buy, "last_use": last_use, "branch": branch
    }

print(f"   오라클 패키지 잔여>0 (2년내): {len(oracle_pkgs)}건")

# 오라클 고객 정보
all_oracle_memnos = list(set(k[0] for k in oracle_pkgs))
oracle_mem = {}
for i in range(0, len(all_oracle_memnos), 500):
    batch = all_oracle_memnos[i:i+500]
    cur.execute(f"SELECT NO, NAME, TEL FROM MEMBER WHERE NO IN ({','.join(str(m) for m in batch)})")
    for r in cur.fetchall():
        oracle_mem[r[0]] = {"name": r[1], "tel": r[2]}

# ═══════════════════════════════════════════
# 2. 오라클 매출메모 기반 선불금(다담권)
# ═══════════════════════════════════════════
print("2. 매출메모 선불금 추출 중...")
cur.execute("""
    SELECT o.MEMNO, m.NAME, m.TEL, o.ORDERDATE, o.NOTE, o.BRANCHNO
    FROM ORDERS o
    JOIN MEMBER m ON o.MEMNO = m.NO
    WHERE o.NOTE LIKE '%잔여%' OR o.NOTE LIKE '%다담%' OR o.NOTE LIKE '%선불%'
    ORDER BY o.MEMNO, o.ORDERDATE DESC
""")
prepaid_rows = cur.fetchall()

memo_prepaid = {}  # key=memno
seen = set()
for memno, name, tel, dt, note, branch in prepaid_rows:
    if memno in seen:
        continue
    seen.add(memno)
    matches = re.findall(r'잔여[금액:\s]*([0-9,.]+)', note or "")
    if matches:
        amt_str = matches[-1].replace(",", "").replace(".", "").strip()
        try:
            amt = int(amt_str)
            if amt > 0 and dt and (today - dt).days <= 730:
                memo_prepaid[memno] = {
                    "name": name, "tel": tel, "amount": amt,
                    "last_date": dt, "branch": branch, "memo": (note or "")[:100]
                }
        except:
            pass

print(f"   매출메모 선불금 잔여>0: {len(memo_prepaid)}명")
cur.close()
conn.close()

# ═══════════════════════════════════════════
# 3. 블리스 Supabase customer_packages
# ═══════════════════════════════════════════
print("3. 블리스 customer_packages 조회 중...")
bliss_pkgs_raw = []
offset = 0
while True:
    r = requests.get(
        f"{SB_URL}/rest/v1/customer_packages?select=*&business_id=eq.biz_khvurgshb&order=id&offset={offset}&limit=1000",
        headers=SB_HEADERS, timeout=15
    )
    batch = r.json() if r.ok else []
    if not batch:
        break
    bliss_pkgs_raw.extend(batch)
    offset += len(batch)
    if len(batch) < 1000:
        break

print(f"   블리스 패키지: {len(bliss_pkgs_raw)}건")

# 블리스 고객 정보 (cust_num → memno 매핑)
print("   블리스 고객 매핑 중...")
bliss_custs = {}
offset = 0
while True:
    r = requests.get(
        f"{SB_URL}/rest/v1/customers?select=id,name,phone,cust_num&business_id=eq.biz_khvurgshb&offset={offset}&limit=1000",
        headers=SB_HEADERS, timeout=15
    )
    batch = r.json() if r.ok else []
    if not batch:
        break
    for c in batch:
        bliss_custs[c["id"]] = c
    offset += len(batch)
    if len(batch) < 1000:
        break

# 블리스 패키지 정리
bliss_pkgs = {}  # key=(cust_num, pkg_type_normalized)
for p in bliss_pkgs_raw:
    cust = bliss_custs.get(p.get("customer_id"), {})
    cust_num = cust.get("cust_num", "")
    svc_name = p.get("service_name", "")
    # 서비스명으로 타입 분류 (package_type 컬럼 없음)
    if "다담" in svc_name or "선불" in svc_name or "잔액" in svc_name:
        pkg_type = "prepaid"
    elif "연간" in svc_name:
        pkg_type = "annual"
    elif "에너지" in svc_name:
        pkg_type = "energy"
    elif "PKG" in svc_name or "패키지" in svc_name or "회권" in svc_name:
        pkg_type = "package"
    elif "할인" in svc_name and "%" in svc_name:
        pkg_type = "discount"
    else:
        pkg_type = "package"  # 기본값
    total = p.get("total_count", 0) or 0
    used = p.get("used_count", 0) or 0
    remain = total - used
    # 다담권: total_count에 충전금액, used_count에 사용금액 저장
    is_prepaid = pkg_type == "prepaid"
    remain_amount = remain if is_prepaid else 0

    bliss_pkgs[p["id"]] = {
        "cust_id": p.get("customer_id"),
        "cust_num": cust_num,
        "cust_name": cust.get("name", ""),
        "cust_phone": cust.get("phone", ""),
        "svc_name": svc_name,
        "pkg_type": pkg_type,
        "total": total, "used": used, "remain": remain,
        "remain_amount": remain_amount,
        "note": p.get("note", ""),
        "created_at": p.get("created_at", ""),
    }

print(f"   블리스 패키지 정리: {len(bliss_pkgs)}건")

# ═══════════════════════════════════════════
# 엑셀 생성
# ═══════════════════════════════════════════
print("엑셀 생성 중...")

hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
hdr_font = Font(bold=True, color="FFFFFF", size=10)
ora_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
bliss_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
memo_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
miss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def write_headers(ws, headers, fills=None):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fills[col-1] if fills else hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')

wb = openpyxl.Workbook()

# ─── 시트1: 오라클 패키지 vs 블리스 ───
ws1 = wb.active
ws1.title = "패키지 비교"

headers1 = [
    "고객번호","이름(오라클)","전화(오라클)",
    "오라클_종류","오라클_구매","오라클_사용","오라클_잔여","오라클_마지막구매","오라클_지점",
    "블리스_서비스명","블리스_총횟수","블리스_사용","블리스_잔여","블리스_메모",
    "상태"
]
write_headers(ws1, headers1)

# 매칭: 오라클 고객번호 → 블리스 cust_num
# 블리스를 cust_num별로 인덱싱
bliss_by_custnum = defaultdict(list)
for pid, bp in bliss_pkgs.items():
    if bp["cust_num"] and bp["pkg_type"] in ("package", "", "energy"):
        bliss_by_custnum[bp["cust_num"]].append(bp)

# 오라클 패키지 종류 → 블리스 서비스명 매핑 (fuzzy)
PKG_NAME_MAP = {
    "토탈": ["토탈", "tot"],
    "왁싱": ["왁싱", "whi", "왁패"],
    "인그로우": ["인그로우", "ing"],
    "미백": ["미백", "화이트"],
    "케어": ["케어"],
    "순환관리": ["순환"],
    "순환관리기본": ["순환기본", "순환관리기본"],
    "힐링": ["힐링"],
}

def match_bliss_pkg(oracle_type, bliss_list):
    keywords = PKG_NAME_MAP.get(oracle_type, [oracle_type.lower()])
    for bp in bliss_list:
        sn = bp["svc_name"].lower()
        for kw in keywords:
            if kw.lower() in sn:
                return bp
    return None

row = 2
matched_bliss_ids = set()
for (memno, pkg_type), od in sorted(oracle_pkgs.items(), key=lambda x: (x[1].get("branch",0), x[0])):
    m = oracle_mem.get(memno, {})
    cust_num_str = str(memno)
    bl_list = bliss_by_custnum.get(cust_num_str, [])
    bl = match_bliss_pkg(pkg_type, bl_list)

    if bl:
        matched_bliss_ids.add(id(bl))
        # 비교
        if od["remain"] == bl["remain"]:
            status = "일치"
        else:
            status = f"불일치(오라클{od['remain']}≠블리스{bl['remain']})"
    else:
        status = "블리스 없음"

    ws1.cell(row=row, column=1, value=memno)
    ws1.cell(row=row, column=2, value=m.get("name",""))
    ws1.cell(row=row, column=3, value=m.get("tel",""))
    ws1.cell(row=row, column=4, value=pkg_type)
    ws1.cell(row=row, column=5, value=od["bought"])
    ws1.cell(row=row, column=6, value=od["used"])
    ws1.cell(row=row, column=7, value=od["remain"])
    ws1.cell(row=row, column=8, value=od["last_buy"].strftime("%Y-%m-%d") if od["last_buy"] else "")
    ws1.cell(row=row, column=9, value=BRANCH_MAP.get(od["branch"], str(od["branch"]) if od["branch"] else ""))
    ws1.cell(row=row, column=10, value=bl["svc_name"] if bl else "")
    ws1.cell(row=row, column=11, value=bl["total"] if bl else "")
    ws1.cell(row=row, column=12, value=bl["used"] if bl else "")
    ws1.cell(row=row, column=13, value=bl["remain"] if bl else "")
    ws1.cell(row=row, column=14, value=bl["note"][:50] if bl else "")
    ws1.cell(row=row, column=15, value=status)

    if "없음" in status:
        ws1.cell(row=row, column=15).fill = miss_fill
    elif "불일치" in status:
        ws1.cell(row=row, column=15).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    row += 1

# 블리스에만 있고 오라클에 없는 패키지
for pid, bp in bliss_pkgs.items():
    if bp["pkg_type"] not in ("package", "", "energy"):
        continue
    if id(bp) in matched_bliss_ids:
        continue
    if bp["remain"] <= 0:
        continue
    ws1.cell(row=row, column=1, value=bp["cust_num"])
    ws1.cell(row=row, column=2, value=bp["cust_name"])
    ws1.cell(row=row, column=3, value=bp["cust_phone"])
    ws1.cell(row=row, column=4, value="")
    ws1.cell(row=row, column=10, value=bp["svc_name"])
    ws1.cell(row=row, column=11, value=bp["total"])
    ws1.cell(row=row, column=12, value=bp["used"])
    ws1.cell(row=row, column=13, value=bp["remain"])
    ws1.cell(row=row, column=14, value=bp["note"][:50])
    ws1.cell(row=row, column=15, value="오라클 없음")
    ws1.cell(row=row, column=15).fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    row += 1

ws1.auto_filter.ref = f"A1:O{row-1}"
for col in range(1, 16):
    ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14

# ─── 시트2: 선불금(다담권) 비교 ───
ws2 = wb.create_sheet("선불금 비교")

headers2 = [
    "고객번호","이름","전화",
    "메모_잔액","메모_날짜","메모_지점","메모원문",
    "블리스_서비스명","블리스_총액","블리스_사용액","블리스_잔액","블리스_메모",
    "상태"
]
write_headers(ws2, headers2)

# 블리스 선불금
bliss_prepaid_by_custnum = defaultdict(list)
for pid, bp in bliss_pkgs.items():
    if bp["pkg_type"] == "prepaid":
        bliss_prepaid_by_custnum[bp["cust_num"]].append(bp)

row2 = 2
matched_prepaid_ids = set()
for memno, mp in sorted(memo_prepaid.items()):
    cust_num_str = str(memno)
    bl_list = bliss_prepaid_by_custnum.get(cust_num_str, [])
    bl = bl_list[0] if bl_list else None

    if bl:
        matched_prepaid_ids.add(id(bl))
        diff = abs(mp["amount"] - bl["remain_amount"])
        if diff == 0:
            status = "일치"
        elif diff < 1000:
            status = f"근사(차이{diff})"
        else:
            status = f"불일치(메모{mp['amount']:,}≠블리스{bl['remain_amount']:,})"
    else:
        status = "블리스 없음"

    ws2.cell(row=row2, column=1, value=memno)
    ws2.cell(row=row2, column=2, value=mp["name"])
    ws2.cell(row=row2, column=3, value=mp["tel"])
    ws2.cell(row=row2, column=4, value=mp["amount"])
    ws2.cell(row=row2, column=5, value=mp["last_date"].strftime("%Y-%m-%d") if mp["last_date"] else "")
    ws2.cell(row=row2, column=6, value=BRANCH_MAP.get(mp["branch"], str(mp["branch"]) if mp["branch"] else ""))
    ws2.cell(row=row2, column=7, value=mp["memo"])
    ws2.cell(row=row2, column=8, value=bl["svc_name"] if bl else "")
    ws2.cell(row=row2, column=9, value=bl["total"] if bl else "")
    ws2.cell(row=row2, column=10, value=bl["used"] if bl else "")
    ws2.cell(row=row2, column=11, value=bl["remain_amount"] if bl else "")
    ws2.cell(row=row2, column=12, value=bl["note"][:50] if bl else "")
    ws2.cell(row=row2, column=13, value=status)

    if "없음" in status:
        ws2.cell(row=row2, column=13).fill = miss_fill
    elif "불일치" in status:
        ws2.cell(row=row2, column=13).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    row2 += 1

# 블리스에만 있는 선불금
for pid, bp in bliss_pkgs.items():
    if bp["pkg_type"] != "prepaid":
        continue
    if id(bp) in matched_prepaid_ids:
        continue
    if bp["remain_amount"] <= 0:
        continue
    ws2.cell(row=row2, column=1, value=bp["cust_num"])
    ws2.cell(row=row2, column=2, value=bp["cust_name"])
    ws2.cell(row=row2, column=3, value=bp["cust_phone"])
    ws2.cell(row=row2, column=8, value=bp["svc_name"])
    ws2.cell(row=row2, column=9, value=bp["total"])
    ws2.cell(row=row2, column=10, value=bp["used"])
    ws2.cell(row=row2, column=11, value=bp["remain_amount"])
    ws2.cell(row=row2, column=12, value=bp["note"][:50])
    ws2.cell(row=row2, column=13, value="메모 없음 (블리스만)")
    ws2.cell(row=row2, column=13).fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    row2 += 1

ws2.auto_filter.ref = f"A1:M{row2-1}"
for col in range(1, 14):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14
ws2.column_dimensions['G'].width = 50

# ─── 시트3: 블리스 전체 목록 ───
ws3 = wb.create_sheet("블리스 전체")
headers3 = ["cust_num","고객명","전화","타입","서비스명","총횟수/총액","사용","잔여","메모","등록일"]
write_headers(ws3, headers3)

row3 = 2
for pid, bp in sorted(bliss_pkgs.items(), key=lambda x: (x[1]["pkg_type"], x[1]["cust_name"])):
    ws3.cell(row=row3, column=1, value=bp["cust_num"])
    ws3.cell(row=row3, column=2, value=bp["cust_name"])
    ws3.cell(row=row3, column=3, value=bp["cust_phone"])
    ws3.cell(row=row3, column=4, value=bp["pkg_type"])
    ws3.cell(row=row3, column=5, value=bp["svc_name"])
    ws3.cell(row=row3, column=6, value=bp["total"])
    ws3.cell(row=row3, column=7, value=bp["used"])
    ws3.cell(row=row3, column=8, value=bp["remain"])
    ws3.cell(row=row3, column=9, value=bp["note"][:80])
    ws3.cell(row=row3, column=10, value=bp["created_at"][:10] if bp["created_at"] else "")
    row3 += 1

ws3.auto_filter.ref = f"A1:J{row3-1}"
for col in range(1, 11):
    ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14

# 저장
outpath = r"C:\Users\TP005\Desktop\패키지_3소스_비교_v2.xlsx"
wb.save(outpath)
print(f"\n저장: {outpath}")

# 통계
oracle_only = sum(1 for (memno, pkg_type), od in oracle_pkgs.items()
                  if not match_bliss_pkg(pkg_type, bliss_by_custnum.get(str(memno), [])))
bliss_only_pkg = sum(1 for pid, bp in bliss_pkgs.items()
                     if bp["pkg_type"] in ("package","") and bp["remain"] > 0 and id(bp) not in matched_bliss_ids)
memo_only = sum(1 for memno in memo_prepaid if str(memno) not in {bp["cust_num"] for bp in bliss_pkgs.values() if bp["pkg_type"]=="prepaid"})

print(f"\n=== 비교 통계 ===")
print(f"오라클 패키지 잔여>0: {len(oracle_pkgs)}건")
print(f"매출메모 선불금 잔여>0: {len(memo_prepaid)}명")
print(f"블리스 전체: {len(bliss_pkgs)}건")
print(f"")
print(f"패키지 - 오라클에만 있음 (블리스 없음): {oracle_only}건")
print(f"패키지 - 블리스에만 있음 (오라클 없음): {bliss_only_pkg}건")
print(f"선불금 - 메모에만 있음 (블리스 없음): {memo_only}명")
